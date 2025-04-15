import requests
from urllib.parse import quote
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as ExcelImage
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import os
import urllib3
import matplotlib.font_manager as fm

urllib3.disable_warnings()

# 인증 및 기본 설정
API_KEY = '206a1f4cf403b9db7af8a982b4951065571d7813'
BASE_URL = 'https://redmine.cresyn.com'
PROJECTS = {
    'hesh_anc2-bnp-skc202stka': 'HESH ANC2',
    'hesh_360-bhp-skc201stka-evo2': 'HESH EVO2',
    'hdx-2990': 'HDX 2990',
    'hdx-3004': 'HDX 3004',
    'ear-x': 'EAR-X'
}
AUTHORS = ['품질보증팀 김예지', '품질보증팀 이효빈', '품질보증팀 이충연']
PRIORITIES = ['A', 'B', 'C', 'D']
STATUSES = ['Open', 'In Progress', 'Resolved', 'Closed']
COLOR_MAP = {'A': '#FF6666', 'B': '#66CC66', 'C': '#FFD700', 'D': '#87CEEB'}
PHASE_MAP = {'EVT': 'EVT', 'EP': 'EVT', 'DVT': 'DVT', 'PP': 'DVT', 'PVT': 'PVT', 'PMP': 'PVT'}

# 버그 커브 생성 함수
def generate_weekly_bug_curve(issues, project_name, project_id):
    plt.rcParams['font.family'] = 'Malgun Gothic'
    plt.rcParams['axes.unicode_minus'] = False
    RESOLVED_STATUSES = {"Resolved", "Closed"}
    NOT_RESOLVED_STATUSES = {"Open", "In Progress"}

    created_counts = defaultdict(int)
    resolved_counts = defaultdict(int)

    for issue in issues:
        created_date = datetime.strptime(issue['created_on'][:10], "%Y-%m-%d").date()
        created_counts[created_date] += 1
        if issue['status']['name'] in RESOLVED_STATUSES:
            resolved_raw = issue.get('closed_on') or issue.get('updated_on')
            if resolved_raw:
                resolved_date = datetime.strptime(resolved_raw[:10], "%Y-%m-%d").date()
                resolved_counts[resolved_date] += 1

    all_dates = sorted(set(created_counts) | set(resolved_counts))
    if not all_dates:
        return None

    df = pd.DataFrame(index=pd.to_datetime(all_dates))
    df['Created'] = pd.Series(created_counts)
    df['Resolved'] = pd.Series(resolved_counts)
    df = df.fillna(0).sort_index()
    df['Total Bug'] = df['Created'].cumsum()
    df['Resolved Total'] = df['Resolved'].cumsum()
    df['Not Resolved'] = df['Total Bug'] - df['Resolved Total']

    plt.figure(figsize=(10, 5))
    plt.plot(df.index, df['Total Bug'], label='Total Bug', color='black', linewidth=2)
    plt.plot(df.index, df['Resolved Total'], label='Resolved', color='dodgerblue', linewidth=2)
    plt.plot(df.index, df['Not Resolved'], label='Not Resolved', color='red', linewidth=2)

    for col, color in zip(['Total Bug', 'Resolved Total', 'Not Resolved'], ['black', 'dodgerblue', 'red']):
        y_val = df[col].iloc[-1]
        x_val = df.index[-1]
        plt.text(x_val, y_val + 1, f"{int(y_val)}", color=color, fontsize=9, ha='center', va='bottom')

    plt.gca().xaxis.set_major_locator(mdates.WeekdayLocator(byweekday=mdates.MO, interval=1))
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%y%m%d'))

    plt.title(f"{project_name} - 누적 이슈 커브", fontsize=14)
    plt.xlabel("날짜")
    plt.ylabel("이슈 수")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.xticks(rotation=45)

    img_path = f"{project_id}_issue_curve.png"
    plt.savefig(img_path)
    plt.close()
    return img_path

# 날짜 기준 설정
today = datetime.today().date()
week_start = today - timedelta(days=today.weekday())
date_str = today.strftime('%Y-%m-%d')

# 워크북 생성
wb = Workbook()
wb.remove(wb.active)
bold = Font(bold=True)

for pid, pname in PROJECTS.items():
    url = f"{BASE_URL}/issues.json?project_id={quote(pid)}&limit=100&status_id=*"
    headers = {'X-Redmine-API-Key': API_KEY}

    try:
        response = requests.get(url, headers=headers, verify=False)
        response.raise_for_status()
        raw_issues = response.json().get('issues', [])
    except Exception as e:
        ws = wb.create_sheet(title=pname[:31])
        ws.append(["프로젝트 API 호출 실패", str(e)])
        continue

    issues = [i for i in raw_issues if i['author']['name'] in AUTHORS]
    ws = wb.create_sheet(title=pname[:31])
    ws.column_dimensions['A'].width = 25
    ws.append([f"SQA 테스트 요약 보고서 ({date_str} 기준)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["[프로젝트명]", pname])
    ws.append(["총 이슈 수", f"{len(issues)}건"])
    not_resolved_count = sum(
        1 for i in issues if i['status']['name'] in ["Open", "In Progress"]
    )
    ws.append(["잔여 이슈 수", f"{not_resolved_count}건"])
    ws.append([])

    # ▶▶▶ [전체 이슈 요약] - 표 형태로 출력
    ws.append(["[전체 이슈 요약]"])
    ws.append([])
    phase_count = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

    for issue in issues:
        subject = issue['subject'].upper()
        phase = '기타'
        for key, value in PHASE_MAP.items():
            if key in subject:
                phase = value
                break
        prio = issue['priority']['name']
        status = issue['status']['name']
        phase_count[phase][prio][status] += 1

    for phase in sorted(phase_count.keys()):
        total = sum(
            count for prio_dict in phase_count[phase].values()
            for count in prio_dict.values()
        )
        ws.append([f"▶ {phase} 단계 ({total}건)"])
        header = ["Priority"] + STATUSES + ["Total"]
        ws.append(header)

        for prio in PRIORITIES:
            row = [prio]
            prio_dict = phase_count[phase].get(prio, {})
            total_by_prio = 0
            for status in STATUSES:
                count = prio_dict.get(status, 0)
                row.append(count)
                total_by_prio += count
            row.append(total_by_prio)
            if total_by_prio > 0:
                ws.append(row)
        ws.append([])

    # ▶▶▶ [이번주 등록 이슈 요약] - 표 형태로 출력
    ws.append(["[이번주 등록 이슈 요약]"])
    weekly_status = defaultdict(lambda: defaultdict(int))

    for issue in issues:
        created = datetime.strptime(issue['created_on'][:10], "%Y-%m-%d").date()
        if created >= week_start:
            prio = issue['priority']['name']
            status = issue['status']['name']
            weekly_status[prio][status] += 1

    if not weekly_status:
        ws.append(["", "▶ 이번 주 등록된 이슈가 없습니다. (0건)"])
    else:
        ws.append(["Priority"] + STATUSES + ["Total"])
        for prio in PRIORITIES:
            row = [prio]
            total = 0
            for status in STATUSES:
                count = weekly_status[prio].get(status, 0)
                row.append(count)
                total += count
            row.append(total)
            if total > 0:
                ws.append(row)
    ws.append([])

    # 버그 커브 이미지 삽입
    curve_img = generate_weekly_bug_curve(issues, pname, pid)
    if curve_img and os.path.exists(curve_img):
        img = ExcelImage(curve_img)
        img.anchor = "I2"
        ws.add_image(img)

# 저장
file_name = f"SQA_주간보고서_{today.strftime('%Y%m%d')}.xlsx"
wb.save(file_name)
print(f"보고서 저장 완료: {file_name}")

# 이미지 삭제
for fname in os.listdir():
    if fname.endswith("_issue_curve.png"):
        try:
            os.remove(fname)
            print(f"{fname} 삭제 완료")
        except Exception as e:
            print(f"삭제 실패: {fname} - {e}")
